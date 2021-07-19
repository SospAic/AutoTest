"""
-------------------------------------------------
  File Name：   Get_Net_sec_protection
  Description :
  Author :    Adonet
  date：     2021/06/28
-------------------------------------------------
  Change Activity:
          2021/07/14:
-------------------------------------------------
"""
__author__ = 'adonet'
import base64
import os
import re
import time
import random
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import xlrd
from xlutils3.copy import copy
from PIL import Image
from pytesseract import pytesseract
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from WebTest.Selenium_driver_base import selenium_driver
from aip import AipOcr
from PIL import Image


class BaiduAip:
    def __init__(self, image_path):
        # API
        self.APP_ID = '24107714'
        self.API_KEY = 'guKhh1ICHk3oFjoxWQZS8jIS'
        self.SECRET_KEY = 'cAvRRxhIO19dONgQSfftTDyRsSOR7N5t'
        # 初始化AipFace对象
        self.client = AipOcr(self.APP_ID, self.API_KEY, self.SECRET_KEY)
        # 可选参数
        self.options = {}
        self.options["language_type"] = "CHN_ENG"  # 中英文混合
        self.options["detect_direction"] = "false"  # 检测朝向
        self.options["detect_language"] = "false"  # 是否检测语言
        self.options["probability"] = "false"  # 是否返回识别结果中每一行的置信度
        self.image_path = image_path

    def get_file_content(self):
        """ 读取图片 """
        with open(self.image_path, 'rb') as fp:
            return fp.read()

    def general_basic(self):
        """通用文字识别"""
        result = self.client.basicGeneral(self.get_file_content(), self.options)
        # 格式化输出-提取需要的部分
        if 'words_result' in result:
            text = ('\n'.join([w['words'] for w in result['words_result']]))
        # """ save """
        # fs = open("baidu_ocr.txt", 'w+')  # 将str,保存到txt
        # fs.write(text)
        # fs.close()
        return text

    def accurate_basic(self):
        """通用文字识别（高精度版）"""
        # img = base64.b64encode(self.get_file_content())
        result = self.client.basicAccurate(self.get_file_content(), self.options)
        if 'words_result' in result:
            text = ('\n'.join([w['words'] for w in result['words_result']]))
        return text


# Log日志记录
class Logger(object):
    def __init__(self, filename="./log/Default.log"):
        self.terminal = sys.stdout
        self.log = open(filename, "a", encoding="utf-8")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)

    def flush(self):
        pass


class Excel:
    def __init__(self, filepath='./output_file/'):
        """ 设置表头单元格及文本样式 """
        self.active_num = 1
        self.data_length = 0
        self.filepath = filepath
        now = datetime.now()
        # create_num = str(random.randint(1, 1000))
        self.create_date = now.strftime('%Y-%m-%d_%H%M%S')  # 创建日期，可配合随机码使用
        self.create_name = '{}定级信息{}.xlsx'.format(self.filepath, self.create_date)
        if os.path.exists(self.create_name):
            # 打开已有表
            self.new = openpyxl.load_workbook(self.create_name)
        else:
            # 创建文件
            self.new = openpyxl.Workbook()
        self.new.remove(self.new.worksheets[0])

    def title_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=15, bold=True, italic=False, vertAlign=None,
                    underline='none', strike=False, color='ffffff')
        fill = PatternFill(fill_type='solid', fgColor='000000')
        border = Border(left=Side(border_style='thick', color='FF000000'),
                        right=Side(border_style='thick', color='FF000000'),
                        top=Side(border_style='thick', color='FF000000'),
                        bottom=Side(border_style='thick', color='FF000000'),
                        diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                        outline=Side(border_style=None, color='FF000000'),
                        vertical=Side(border_style=None, color='FF000000'),
                        horizontal=Side(border_style=None, color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def content_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=11, bold=True, italic=False, vertAlign=None,
                    underline='none', strike=False, color='ffffff')
        fill = PatternFill(fill_type='solid', fgColor='1e90ff')
        border = Border(left=Side(border_style='thick', color='FF000000'),
                        right=Side(border_style='thick', color='FF000000'),
                        top=Side(border_style='thick', color='FF000000'),
                        bottom=Side(border_style='thick', color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def detail_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=11, bold=False, italic=False, vertAlign=None,
                    underline='none', strike=False, color='000000')
        fill = PatternFill(fill_type='solid', fgColor='40e0d0')
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def index_create(self, sheet_title=None, content_title=None, content=None):
        """ 创建目录 """
        row = 1
        # 创建Sheet页
        if len(self.new.worksheets) == 0:
            self.new.create_sheet(sheet_title)
            print('创建目录')
        else:
            print('追加目录')
        self.new._active_sheet_index = 0  # 改变活动sheet,=后面的数字代表sheet在工作薄的排序号
        ws = self.new[self.new.sheetnames[0]]
        if ws.max_row <= 1:
            # 新建表头
            for col in range(1, len(content_title) + 1):
                write_data = content_title[col - 1]
                self.title_style_write(1, col, write_data)
                print('列名共{0:>2}条,正在添加：{1}'.format(str(len(content_title)), write_data))
        else:
            # 追加目录
            row = ws.max_row
        col = 1
        # 写入内容
        for total in range(len(content)):
            write_data = content[total]
            total += 1
            print('共计{0:>3}条信息，当前为{1:>3}条：{2}'.format(len(content), total, write_data))
            if re.match(r'^\d{1,3}$', write_data):
                row += 1
                col = 1
            else:
                pass
            self.content_style_write(row, col, write_data)
            col += 1
        # 保存表格
        try:
            self.new.save(self.create_name)
            print('目录创建成功，当前时间为： ' + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
        except PermissionError:
            print('文件未关闭，保存失败，请关闭文件后重试')
            pass

    def excel_create(self, content_title, content):
        """ 创建详情 """
        # 创建Sheet页
        self.new.create_sheet(content_title)
        self.new._active_sheet_index = self.active_num
        print('创建{}详情'.format(content_title))
        # 写入表头
        self.title_style_write(1, 1, content_title)
        # 写入内容
        row = 2
        total = 1
        sheet = self.new.active
        self.active_num += 1
        row_length = 0
        for index in content:
            print('共计{0:>3}条信息，当前为{1:>3}条：{2}'.format(len(content), total, index))
            col = 1
            # 识别标题
            if type(index) == str:
                self.content_style_write(row, col, index)
                col += 1
            # 识别详细数据
            elif type(index) == list:
                # 将子标题样式合并至与下方等长
                if type(index[0]) == str:
                    row_length = len(index)
                    sheet.merge_cells(
                        '{}{}:{}{}'.format(get_column_letter(1), row - 1, get_column_letter(len(index)), row - 1))
                for aa in index:
                    # 子标题
                    if type(aa) == str:
                        self.content_style_write(row, col, aa)
                        col += 1
                    # 详细信息
                    elif type(aa) == list:
                        temp = len(aa)
                        for a in aa:
                            self.detail_style_write(row, col, a)
                            # 过滤“没有数据”
                            if re.match(r'\u6ca1\u6709\u6570\u636e', a):
                                sheet.merge_cells(
                                    '{}{}:{}{}'.format(get_column_letter(1), row, get_column_letter(len(a) + 1), row))
                            else:
                                pass
                            col += 1
                        try:
                            # 空位填充
                            if row_length > temp:
                                circle = row_length - temp
                                # print('row_length:{},temp:{},circle:{}'.format(row_length, temp, circle))
                                for b in range(circle):
                                    self.detail_style_write(row, col, '')
                                    col += 1
                            elif row_length < temp:
                                row_length = temp
                        except AttributeError:
                            pass
                        if aa != index[-1]:
                            row += 1
                        else:
                            pass
                        col = 1
            total += 1
            row += 1
        # 保存表格
        try:
            self.new.save(self.create_name)
            print('保存表格成功，当前时间为： ' + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
        except PermissionError:
            print('文件未关闭，保存失败，请关闭文件后重试')
            pass

    def auto_set_row_col(self):
        # 自动设置所有表格的列宽和行高
        wb = openpyxl.load_workbook(self.create_name)
        # print(wb.sheetnames)
        # 设置行高、列宽
        # 遍历所有Sheet页
        for a in range(len(wb.sheetnames)):
            ws = wb[wb.sheetnames[a]]
            self.new._active_sheet_index = a
            wd = self.new.active
            # print("row:", ws.max_row, "column:", ws.max_column)
            if a == 0:
                ws.merge_cells('D1:H1')
            else:
                ws.merge_cells('A1:I1')
            ws.row_dimensions[1].height = 22
            for i in range(1, ws.max_column + 1):
                column_data = []
                # 遍历列中所有单元格长度
                for ii in range(1, ws.max_row + 1):
                    cell_value = wd.cell(row=ii, column=i).value
                    a_length = len(str(cell_value).encode('utf-8'))
                    # print('a:{},i:{},ii:{},cell_value:{},a_length:{}'.format(a, i, ii, cell_value, a_length))
                    column_data.append(a_length)
                    # ws.row_dimensions[i].height = height
                col_width = max(column_data)
                ws.column_dimensions[get_column_letter(i)].width = col_width
        wb.save(self.create_name)

    def set_row_col(self):
        # 预设特定表格的列宽和行高
        wb = openpyxl.load_workbook(self.create_name)
        for a in range(len(wb.sheetnames)):
            ws = wb[wb.sheetnames[a]]
            if a == 0:
                ws.merge_cells('D1:H1')
                ws.row_dimensions[1].height = 22
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 11
                ws.column_dimensions['E'].width = 6
                ws.column_dimensions['F'].width = 7
                ws.column_dimensions['G'].width = 13
                ws.column_dimensions['H'].width = 15
            else:
                ws.merge_cells('A1:I1')
                ws.row_dimensions[1].height = 22
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 11
                ws.column_dimensions['D'].width = 29
                ws.column_dimensions['E'].width = 11
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 23
                ws.column_dimensions['H'].width = 17
                ws.column_dimensions['I'].width = 9
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 5
        wb.save(self.create_name)


class GetInfo:
    def __init__(self, url):
        # 忽略证书错误
        options = webdriver.ChromeOptions()
        options.add_argument('--ignore-certificate-errors')
        self.driver = webdriver.Chrome(executable_path=selenium_driver(), chrome_options=options)
        self.driver.implicitly_wait(30)
        self.driver.get(url)
        self.driver.maximize_window()
        self.login_name = "***"
        self.password = "***"
        self.login_num = 1
        self.screenshot_file = './pic_code/auth_cap.png'
        self.verify_file = './pic_code/auth_cap.png'
        self.verify_code = ''
        self.excel = Excel()
        self.per_page_num = '10'
        self.turn = True

    def is_exist_element(self, elem, code='错误信息'):
        """判断元素是否存在"""
        try:
            s = self.driver.find_element_by_xpath(elem)
            if s.text == code:
                return True
            elif s:
                return True
            else:
                return False
        except NoSuchElementException:
            return False

    def get_auth_code(self):
        """获取验证码"""
        self.driver.save_screenshot(self.screenshot_file)  # 截取登录页面
        img_element = self.driver.find_element_by_id('identifyCodeImg')
        img_size = img_element.size  # 获取验证码图片的大小
        img_location = img_element.location  # 获取验证码元素坐标
        img_location['x'] = img_location['x'] + 305  # 偏移量
        img_location['y'] = img_location['y'] + 85
        rangle = (int(img_location['x']), int(img_location['y']),
                  int(img_location['x'] + img_size['width']),
                  int(img_location['y'] + img_size['height']))  # 计算验证码整体坐标
        login = Image.open(self.screenshot_file)
        frame4 = login.crop(rangle)  # 截取验证码图片
        frame4.save(self.verify_file)
        # auth_code_img = Image.open(self.verify_file)
        # self.auth_code_text = pytesseract.image_to_string(auth_code_img).strip()  # tesseract识别
        # return self.auth_code_text
        baidu_ocr = BaiduAip(self.verify_file)  # baidu_ocr识别
        code_result = baidu_ocr.general_basic()
        self.verify_code = code_result
        return code_result

    def login(self):
        """OCR识别登录"""
        while self.login_num <= 100:
            time.sleep(0.5)
            self.get_auth_code()
            self.driver.find_element(By.ID, "LOGIN_NAME").send_keys(self.login_name)
            self.driver.find_element(By.ID, "PWD").send_keys(self.password)
            if self.verify_code != '':
                self.driver.find_element(By.ID, "email").send_keys(self.verify_code)
            else:
                self.driver.refresh()
                self.login()
            self.driver.find_element(By.ID, "btnLogin").click()
            if self.is_exist_element('//*[@id="btnLogin"]'):
                self.login_num += 1
                time.sleep(0.5)
                self.login()
            else:
                # 登陆成功
                self.get_start()
                break

    def normal_login(self):
        """手动登录"""
        while self.login_num <= 100:
            time.sleep(0.5)
            self.driver.find_element(By.ID, "LOGIN_NAME").send_keys(self.login_name)
            self.driver.find_element(By.ID, "PWD").send_keys(self.password)
            self.driver.find_element(By.ID, "email").click()
            time.sleep(4)
            self.driver.find_element(By.ID, "btnLogin").click()
            if self.is_exist_element('//*[@id="btnLogin"]'):
                time.sleep(0.5)
                self.driver.find_element(By.ID, "LOGIN_NAME").clear()
                self.driver.find_element(By.ID, "PWD").clear()
                self.normal_login()
            else:
                # 登陆成功
                self.get_start()
                break

    def title_base(self):
        title_list = []
        data_result = []
        data_base = self.driver.find_elements(By.ID, "_GRADINGOBJ_tblDatasTable")
        # 数据标题
        data_title = data_base[0].find_elements(By.XPATH, "./thead/tr/th")
        for i in data_title:
            if i == '':
                data_title.remove(i)
                pass
            else:
                title_list.append(i.text)
        title_list.remove('')
        print(title_list)
        # 基础数据
        data_lists = data_base[0].find_elements(By.XPATH, "//tbody/tr/*")
        # 数据清洗
        for i in data_lists:
            i = i.text
            if re.match(r'^[\u67e5\u8be2\u7ed3\u679c]+', i):
                pass
            elif re.match(r'^[\u8be6\u7ec6]+', i):
                pass
            elif re.match(r"^\d{4}", i):
                data_result.append(i)
            elif re.match(r'^\| [\u4e00-\u9fa5]*', i):
                a = i.split("|")
                # c = list(filter(None, a))
                for b in a:
                    b = b.strip()
                    if b == '':
                        pass
                    else:
                        data_result.append(b)
                pass
            elif i == '':
                pass
            else:
                data_result.append(i)
        print(data_result)
        self.excel.index_create(sheet_title='目 录', content_title=title_list, content=data_result)

    def content_base(self, run_num, data_details):
        # 详情信息
        for i in range(run_num):
            try:
                contents = []
                # time.sleep(2)
                data_details = self.driver.find_elements(By.XPATH,
                                                         '//*[@id="_GRADINGOBJ_tblDatasTable"]/tbody/tr[@role]')
                main_device_title = []
                detail_name = data_details[i].find_element(By.XPATH, "./td[3]").text
                print('第{}条，正在获取<{}>详情'.format(i + 1, detail_name))
                time.sleep(5)
                try:
                    detail_page = data_details[i].find_element(By.XPATH, "./td[5]/a[1]").click()
                except StaleElementReferenceException:
                    until_run = True
                    while until_run:
                        time.sleep(1)
                        print('元素不可见，1秒后重试')
                        if self.is_exist_element(elem="./td[5]/a[1]"):
                            data_details[i].find_element(By.XPATH, "./td[5]/a[1]").click()
                            until_run = False
                        elif self.is_exist_element(elem="//a[contains(text(),'返回')]"):
                            self.driver.find_element(By.XPATH, '//*[@id="_page-buttons"]/a'). \
                                find_element(By.XPATH, "//a[contains(text(),'返回')]").click()
                            time.sleep(2)
                            if self.is_exist_element(elem="./td[5]/a[1]"):
                                data_details[i].find_element(By.XPATH, "./td[5]/a[1]").click()
                                until_run = False
                            else:
                                time.sleep(1)
                                continue
                        else:
                            time.sleep(1)
                            continue
                time.sleep(3)
                # 主要硬件
                main_device = self.driver.find_elements(By.XPATH,
                                                        '//*[@id="_GRADINGOBJINFO_tblhardwareTable"]/thead/tr'
                                                        '/th')
                main_device_title_name = self.driver.find_element(By.XPATH, '//*[@id="pnlhardware"]').find_element(
                    By.TAG_NAME, 'label').text
                print(main_device_title_name)
                for j in main_device:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        main_device_title.append(j)
                # main_device_title.remove('')
                detail_list = self.driver.find_elements(By.XPATH,
                                                        '//*[@id="_GRADINGOBJINFO_tblhardwareTable"]/tbody/tr')
                device_list = []
                for a in detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    device_list.append(every_list)
                print(main_device_title)
                print(device_list)
                # 安全系统
                safe_sys_title = []
                safe_sys = self.driver.find_elements(By.XPATH,
                                                     '//*[@id="_GRADINGOBJINFO_tblsafesysTable"]/thead/tr/th')
                safe_sys_title_name = self.driver.find_element(By.XPATH, '//*[@id="pnlsafesys"]').find_element(
                    By.TAG_NAME,
                    'label').text
                print(safe_sys_title_name)
                for j in safe_sys:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        safe_sys_title.append(j)
                # safe_sys_title.remove('')
                safe_sys_detail_list = self.driver.find_elements(By.XPATH,
                                                                 '//*[@id="_GRADINGOBJINFO_tblsafesysTable"]/tbody/tr')
                safe_sys_device_list = []
                for a in safe_sys_detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    safe_sys_device_list.append(every_list)
                print(safe_sys_title)
                print(safe_sys_device_list)
                # 基础软件
                basic_software_title = []
                basic_software = self.driver.find_elements(By.XPATH,
                                                           '//*[@id="_GRADINGOBJINFO_tblsoftwareTable"]/thead/tr/th')
                basic_software_title_name = self.driver.find_element(By.XPATH,
                                                                     '//*[@id="pnlsoftware"]').find_element(
                    By.TAG_NAME, 'label').text
                print(basic_software_title_name)
                for j in basic_software:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        basic_software_title.append(j)
                # basic_software_title.remove('')
                basic_software_detail_list = self.driver.find_elements(By.XPATH,
                                                                       '//*[@id="_GRADINGOBJINFO_tblsoftwareTable"]/tbody/tr')
                basic_software_device_list = []
                for a in basic_software_detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    basic_software_device_list.append(every_list)
                print(basic_software_title)
                print(basic_software_device_list)
                # 应用软件
                application_software_title = []
                application_software = self.driver.find_elements(By.XPATH,
                                                                 '//*[@id="_GRADINGOBJINFO_tblappTable"]/thead/tr'
                                                                 '/th')
                application_software_title_name = self.driver.find_element(By.XPATH,
                                                                           '//*[@id="pnlapp"]').find_element(
                    By.TAG_NAME, 'label').text
                print(application_software_title_name)
                for j in application_software:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        application_software_title.append(j)
                # application_software_title.remove('')
                application_software_detail_list = self.driver.find_elements(By.XPATH,
                                                                             '//*[@id="_GRADINGOBJINFO_tblappTable"]/tbody/tr')
                application_software_device_list = []
                for a in application_software_detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    application_software_device_list.append(every_list)
                print(application_software_title)
                print(application_software_device_list)
                # 公网IP地址段
                ip_table_title = []
                ip_table = self.driver.find_elements(By.XPATH, '//*[@id="_GRADINGOBJINFO_tblIPTable"]/thead/tr'
                                                               '/th')
                ip_table_title_name = self.driver.find_element(By.XPATH, '//*[@id="pnlIP"]').find_element(
                    By.TAG_NAME,
                    'label').text
                print(ip_table_title_name)
                for j in ip_table:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        ip_table_title.append(j)
                # ip_table_title.remove('')
                ip_table_detail_list = self.driver.find_elements(By.XPATH,
                                                                 '//*[@id="_GRADINGOBJINFO_tblIPTable"]/tbody/tr')
                ip_table_device_list = []
                for a in ip_table_detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    try:
                        every_list.remove('修改添加端口')
                    except ValueError:
                        pass
                    ip_table_device_list.append(every_list)
                print(ip_table_title)
                print(ip_table_device_list)
                # 主要协议和端口
                port_table_title = []
                port_table = self.driver.find_elements(By.XPATH, '//*[@id="_GRADINGOBJINFO_tblPORTTable"]/thead/tr'
                                                                 '/th')
                port_table_title_name = self.driver.find_element(By.XPATH, '//*[@id="pnlPORT"]').find_element(
                    By.TAG_NAME,
                    'label').text
                print(port_table_title_name)
                for j in port_table:
                    j = j.text
                    if j == '':
                        pass
                    else:
                        port_table_title.append(j)
                # port_table_title.remove('')
                port_table_detail_list = self.driver.find_elements(By.XPATH,
                                                                   '//*[@id="_GRADINGOBJINFO_tblPORTTable"]/tbody/tr')
                port_table_device_list = []
                for a in port_table_detail_list:
                    every_detail = a.find_elements(By.XPATH, './td')
                    every_list = []
                    for m in every_detail:
                        m = m.text
                        if m == '':
                            pass
                        elif re.match(r'^\| +[\u4e00-\u9fa5]*', m):
                            a = m.split("|")
                            for b in a:
                                b = b.strip()
                                if b == '':
                                    pass
                                else:
                                    every_list.append(b)
                        else:
                            every_list.append(m)
                    port_table_device_list.append(every_list)
                print(port_table_title)
                print(port_table_device_list)
                self.driver.find_element(By.XPATH, '//*[@id="_page-buttons"]/a'). \
                    find_element(By.XPATH, "//a[contains(text(),'返回')]").click()
                if i == run_num:
                    print('当前页数据获取完成')
                    break
                else:
                    print('返回上一页')
                final_list = [main_device_title_name, main_device_title, device_list, safe_sys_title_name,
                              safe_sys_title,
                              safe_sys_device_list, basic_software_title_name, basic_software_title,
                              basic_software_device_list,
                              application_software_title_name, application_software_title,
                              application_software_device_list,
                              ip_table_title_name, ip_table_title, ip_table_device_list, port_table_title_name,
                              port_table_title,
                              port_table_device_list]
                for item in final_list:
                    contents.append(item)
                self.excel.excel_create(content_title=detail_name, content=contents)
            except IndexError:
                length_check = len(self.driver.find_elements(By.XPATH,
                                                             '//*[@id="_GRADINGOBJ_tblDatasTable"]/tbody/tr[@role]'))
                # print('i:{}, length_check:{}'.format(i, length_check))
                if i == length_check:
                    print('所有数据已获取完毕，正在处理...')
                    time.sleep(2)
                    self.driver.quit()
                    self.excel.set_row_col()
                    self.turn = False
                    return 0
                else:
                    print('获取Index失败，继续下一条')
                    self.driver.refresh()
                    self.search_options()
                    continue
            except NoSuchElementException:
                print('获取元素失败，继续下一条')
                try:
                    self.driver.find_element(By.XPATH, '//*[@id="_page-buttons"]/a'). \
                        find_element(By.XPATH, "//a[contains(text(),'返回')]").click()
                    continue
                except NoSuchElementException:
                    self.driver.refresh()
                    self.search_options()
                    continue
            except TimeoutException:
                print('请求超时，继续下一条')
                self.driver.refresh()
                self.search_options()
                continue
            except ElementClickInterceptedException:
                print('元素不可点击，1秒后自动重试')
                time.sleep(1)
                try:
                    data_details[i].find_element(By.XPATH, "./td[5]/a[1]").click()
                except ElementClickInterceptedException:
                    print('元素不可点击，1秒后自动重试')
                    time.sleep(1)

    def get_start(self):
        self.driver.find_element(By.CSS_SELECTOR, ".\\_top_menu:nth-child(2) > p").click()
        # 检索条件
        self.search_options()
        select_ele = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div/div/div[1]/div/div/div/div["
                                                        "3]/div/div/div/div/div/div/div[2]/div/div/div/div["
                                                        "1]/div/label/select")
        s = Select(select_ele)
        self.per_page_num = '100'
        s.select_by_value(self.per_page_num)
        time.sleep(2)
        data_details = self.driver.find_elements(By.XPATH, '//*[@id="_GRADINGOBJ_tblDatasTable"]/tbody/tr[@role]')
        run_num = len(data_details)
        select_page = len(self.driver.find_elements(By.XPATH, '//*[@id="_GRADINGOBJ_tblDatasTable_paginate"]/ul/li'))
        for page in range(1, select_page):
            self.title_base()
            self.content_base(run_num=run_num, data_details=data_details)
            if int(self.per_page_num) > run_num:
                print('所有数据已获取完毕，正在处理...')
                time.sleep(2)
                self.driver.quit()
                self.excel.set_row_col()
                self.turn = False
                return 0
            else:
                if not self.turn:
                    print('处理完毕，正在退出...')
                    return 0
                else:
                    print('第{}页完成，正在跳转下一页'.format(page))
                    next_button = self.driver.find_element(By.XPATH, '//*[@id="_GRADINGOBJ_tblDatasTable_next"]/a')
                    try:
                        next_button.send_keys(Keys.ENTER)  # click方法无效时可使用enter按键尝试
                        time.sleep(3)
                    except ElementClickInterceptedException:
                        print('下一页按钮不可见，请检查脚本')
                        pass

    def search_options(self):
        # 检索条件
        try:
            time.sleep(3)
            self.driver.find_element(By.CSS_SELECTOR, "#RANK_NO > .checkbox:nth-child(1)").click()  # 1级
            self.driver.find_element(By.CSS_SELECTOR, "#RANK_NO > .checkbox:nth-child(2)").click()  # 2级
            self.driver.find_element(By.CSS_SELECTOR, "#RANK_NO > .checkbox:nth-child(3)").click()  # 3级
            self.driver.find_element(By.CSS_SELECTOR, "#STATE > .checkbox:nth-child(1)").click()  # 未提交
            self.driver.find_element(By.CSS_SELECTOR, "#STATE > .checkbox:nth-child(2)").click()  # 已提交
            self.driver.find_element(By.CSS_SELECTOR, "#STATE > .checkbox:nth-child(3)").click()  # 审核通过
            self.driver.find_element(By.CSS_SELECTOR, "#STATE > .checkbox:nth-child(4)").click()  # 审核未通过
            self.driver.find_element(By.CSS_SELECTOR, "#ORDER > .radio:nth-child(1)").click()
            self.driver.find_element(By.CSS_SELECTOR, "#search > span").click()
            time.sleep(2)
        except ElementClickInterceptedException:
            time.sleep(1)
            self.search_options()
            print('元素不可点击，1秒后自动重试')


if __name__ == '__main__':
    sys.stdout = Logger('./log/抓取网络安全管理日志.log')
    pytesseract.tesseract_cmd = r'./other/tesseract.exe'
    start = GetInfo('https://www.mii-aqfh.cn/')
    # start.login()
    start.normal_login()
