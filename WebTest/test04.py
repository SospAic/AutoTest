import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

sheet = wb.active

# 设置行高

sheet['A1'] = '行高被设置为 100'

sheet.row_dimensions[1].height = 100

# 设置列宽

sheet['B2'] = '列宽被设置为 50'

sheet.column_dimensions['B'].width = 50
print(get_column_letter(3))
# xy = coordinate_from_string('A4') # returns ('A',4)
# col = column_index_from_string(xy[0]) # returns 1
# row = xy[1]
print(sheet.cell(3, 2).column)

# sheet.merge_cells('{}{}:{}{}'.format(1,2))
wb.save('dimensions.xlsx')