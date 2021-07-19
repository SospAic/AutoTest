# -*- coding:utf-8 -*-
import re

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
import os
import time
from datetime import datetime

index_name = ['No.', '定级对象名称', '创建时间', '控制']
index_list = ['1', '“龙江先锋”党建平台', '2021/06/21', '黑龙江联通', '第1级', '未提交', '2021/06/21', '省网络管理中心', '2',
              '中国联通IPRAN光传送网黑龙江省哈尔滨市本地传送网',
              '2019/07/19', '黑龙江联通', '第2级', '已提交', '2019/07/22', '哈尔滨', '3', '中国联通IP承载网黑龙江省哈尔滨市IP城域网核心层', '2010/04/19',
              '黑龙江联通',
              '第3级', '已提交', '2021/02/23', '哈尔滨', '4', '中国联通IP承载网黑龙江省哈尔滨市IP城域网汇聚层', '2010/04/19', '黑龙江联通', '第2级', '已提交',
              '2021/02/23', '哈尔滨', '5', '中国联通IP承载网黑龙江省鸡西市接入网', '2010/04/19', '黑龙江联通', '第2级', '已提交', '2021/02/23', '鸡西',
              '6',
              '中国联通IP承载网黑龙江省鹤岗市IP城域网核心层', '2010/04/19', '黑龙江联通', '第2级', '已提交', '2021/02/23', '鹤岗', '7',
              '中国联通IP承载网黑龙江省齐齐哈尔市IDC',
              '2018/12/21', '黑龙江联通', '第1级', '已提交', '2021/04/29', '齐齐哈尔', '8', '中国联通ip承载网黑龙江省七台河市ip城域网', '2010/04/19',
              '黑龙江联通',
              '第2级', '已提交', '2021/03/26', '七台河', '9', '中国联通ip承载网黑龙江省佳木斯市ip城域网', '2010/04/19', '黑龙江联通', '第2级', '已提交',
              '2021/03/26', '佳木斯', '10', '中国联通ip承载网黑龙江省大兴安岭地区ip城域网', '2010/04/19', '黑龙江联通', '第2级', '已提交', '2021/03/26',
              '大兴安岭']
index_list1 = ['1', '中国联通黑河市本地网综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '黑河', '申请中', '2',
               '中国联通黑龙江省七台河市本地网综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '七台河', '申请中', '3',
               '中国联通黑龙江省分公司双鸭山市本地网综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2019/12/09', '双鸭山', '申请中', '4',
               '中国联通黑龙江省大庆市本地网综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '大庆', '5',
               '中国联通黑龙江省鹤岗市本地网综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '鹤岗', '申请中', '6',
               '中国联通黑龙江省齐齐哈尔市本地网综合关口局', '2015/07/12', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '齐齐哈尔', '7',
               '黑龙江省分公司大兴安岭地区综合关口局', '2010/04/19', '黑龙江联通', '第3级', '审核通过', '2020/01/08', '大兴安岭']
excel_name = 'test'
excel_list = [
    '主要硬件',
    ['No.', '类型', '品牌', '型号', '数量（个）', '维护方式', '所在机房', '已运行时间（天）', '控制'],
    [['1', 'PTN设备', '华为', 'CX600系列、ATN系列', '3334', '远程', '哈尔滨市区', '1781'], ['存在问题 ： 未发现问题', '问题详细 ：']],
    '安全系统',
    ['No.', '品牌/研发方', '型号', '数量（个）', '维护方式', '所在机房', '已运行时间（天）', '控制'],
    [['1', '无', '无', '0', '远程', '无', '0'], ['存在问题 ： 未发现问题', '问题详细 ：']],
    '基础软件',
    ['No.', '类型', '厂商', '当前版本', '套数', '维护方式', '所在机房', '已运行时间（天）', '所在设备', '是否为正版软件', '控制'],
    [['1', '操作系统', '微软', 'win7', '2', '本地', '哈尔滨市尚志局4楼网管', '1781', '联想电脑', '是'], ['存在问题 ：', '问题详细 ：']],
    '应用软件',
    ['No.', '名称', '研发方', '当前版本', '套数', '维护方式', '所在机房', '已运行时间（天）', '所在设备', '控制'],
    [['1', '网管u2000软件', '第三方研发', 'U2000 V200R018C50CP2003', '2', '本地', '哈尔滨尚志局网管4楼', '1781', '联想电脑'],
     ['存在问题 ：', '问题详细 ：']],
    '公网IP地址段',
    ['No.', 'IP分类', 'IP地址段', 'IP地址描述', '控制'],
    [['没有数据']],
    '主要协议和端口',
    ['No.', '端口', '端口描述', '对应IP', '控制'],
    [['没有数据']]
]


class Excel:
    def __init__(self, filepath='./output_file/'):
        """ 设置表头单元格及文本样式 """
        self.active_num = 1
        self.data_length = 0
        self.filepath = filepath
        now = datetime.now()
        # create_num = str(random.randint(1, 1000))
        self.create_date = now.strftime('%Y-%m-%d_%H%M%S')  # 创建日期，可配合随机码使用
        self.create_name = '{}定级信息{}.xlsx'.format(self.filepath, self.create_date)
        if os.path.exists(self.create_name):
            # 打开已有表
            self.new = openpyxl.load_workbook(self.create_name)
        else:
            # 创建文件
            self.new = openpyxl.Workbook()
        self.new.remove(self.new.worksheets[0])

    def title_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=15, bold=True, italic=False, vertAlign=None,
                    underline='none', strike=False, color='ffffff')
        fill = PatternFill(fill_type='solid', fgColor='000000')
        border = Border(left=Side(border_style='thick', color='FF000000'),
                        right=Side(border_style='thick', color='FF000000'),
                        top=Side(border_style='thick', color='FF000000'),
                        bottom=Side(border_style='thick', color='FF000000'),
                        diagonal=Side(border_style=None, color='FF000000'), diagonal_direction=0,
                        outline=Side(border_style=None, color='FF000000'),
                        vertical=Side(border_style=None, color='FF000000'),
                        horizontal=Side(border_style=None, color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def content_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=11, bold=True, italic=False, vertAlign=None,
                    underline='none', strike=False, color='ffffff')
        fill = PatternFill(fill_type='solid', fgColor='1e90ff')
        border = Border(left=Side(border_style='thick', color='FF000000'),
                        right=Side(border_style='thick', color='FF000000'),
                        top=Side(border_style='thick', color='FF000000'),
                        bottom=Side(border_style='thick', color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def detail_style_write(self, row, col, value):
        sheet = self.new.active
        font = Font(name='微软雅黑', size=11, bold=False, italic=False, vertAlign=None,
                    underline='none', strike=False, color='000000')
        fill = PatternFill(fill_type='solid', fgColor='40e0d0')
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center', vertical='bottom', text_rotation=0,
                              wrap_text=False, shrink_to_fit=False, indent=0)
        number_format = 'General'
        protection = Protection(locked=False, hidden=False)
        sheet.cell(row=row, column=col).font = font
        sheet.cell(row=row, column=col).fill = fill
        sheet.cell(row=row, column=col).border = border
        sheet.cell(row=row, column=col).alignment = alignment
        sheet.cell(row=row, column=col).number_format = number_format
        sheet.cell(row=row, column=col).protection = protection
        try:
            sheet.cell(row=row, column=col).value = value
        except:
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            content = ILLEGAL_CHARACTERS_RE.sub(r'', value)
            sheet.cell(row=row, column=col).value = content

    def index_create(self, sheet_title=None, content_title=None, content=None):
        """ 创建目录 """
        row = 1
        # 创建Sheet页
        if len(self.new.worksheets) == 0:
            self.new.create_sheet(sheet_title)
            print('创建目录')
        else:
            print('追加目录')
        self.new._active_sheet_index = 0  # 改变活动sheet,=后面的数字代表sheet在工作薄的排序号
        ws = self.new[self.new.sheetnames[0]]
        if ws.max_row <= 1:
            # 新建表头
            for col in range(1, len(content_title) + 1):
                write_data = content_title[col - 1]
                self.title_style_write(1, col, write_data)
                print('列名共{0:>2}条,正在添加：{1}'.format(str(len(content_title)), write_data))
        else:
            # 追加目录
            row = ws.max_row
        col = 1
        # 写入内容
        for total in range(len(content)):
            write_data = content[total]
            total += 1
            print('共计{0:>3}条信息，当前为{1:>3}条：{2}'.format(len(content), total, write_data))
            if col >= 9:
                if re.match(r'\u7533\u8bf7\u4e2d', write_data):
                    pass
                else:
                    row += 1
                    col = 1
            self.content_style_write(row + 1, col, write_data)
            col += 1
        # 保存表格
        try:
            self.new.save(self.create_name)
            print('目录创建成功，当前时间为： ' + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
        except PermissionError:
            print('文件未关闭，保存失败，请关闭文件后重试')
            pass

    def excel_create(self, content_title, content):
        """ 创建详情 """
        # 创建Sheet页
        self.new.create_sheet(content_title)
        self.new._active_sheet_index = self.active_num
        print('创建{}详情'.format(content_title))
        # 写入表头
        self.title_style_write(1, 1, content_title)
        # 写入内容
        row = 2
        total = 1
        sheet = self.new.active
        self.active_num += 1
        row_length = 0
        for index in content:
            print('共计{0:>3}条信息，当前为{1:>3}条：{2}'.format(len(content), total, index))
            col = 1
            # 识别标题
            if type(index) == str:
                self.content_style_write(row, col, index)
                col += 1
            # 识别详细数据
            elif type(index) == list:
                # 将子标题样式合并至与下方等长
                if type(index[0]) == str:
                    row_length = len(index)
                    sheet.merge_cells(
                        '{}{}:{}{}'.format(get_column_letter(1), row - 1, get_column_letter(len(index)), row - 1))
                for aa in index:
                    # 子标题
                    if type(aa) == str:
                        self.content_style_write(row, col, aa)
                        col += 1
                    # 详细信息
                    elif type(aa) == list:
                        temp = len(aa)
                        for a in aa:
                            self.detail_style_write(row, col, a)
                            # 过滤“没有数据”
                            if re.match(r'\u6ca1\u6709\u6570\u636e', a):
                                sheet.merge_cells(
                                    '{}{}:{}{}'.format(get_column_letter(1), row, get_column_letter(len(a) + 1), row))
                            else:
                                pass
                            col += 1
                        try:
                            # 空位填充
                            if row_length > temp:
                                circle = row_length - temp
                                # print('row_length:{},temp:{},circle:{}'.format(row_length, temp, circle))
                                for b in range(circle):
                                    self.detail_style_write(row, col, '')
                                    col += 1
                            elif row_length < temp:
                                row_length = temp
                        except AttributeError:
                            pass
                        if aa != index[-1]:
                            row += 1
                        else:
                            pass
                        col = 1
            total += 1
            row += 1
        # 保存表格
        try:
            self.new.save(self.create_name)
            print('保存表格成功，当前时间为： ' + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
        except PermissionError:
            print('文件未关闭，保存失败，请关闭文件后重试')
            pass

    def auto_set_row_col(self):
        # 自动设置所有表格的列宽和行高
        wb = openpyxl.load_workbook(self.create_name)
        # print(wb.sheetnames)
        # 设置行高、列宽
        # 遍历所有Sheet页
        for a in range(len(wb.sheetnames)):
            ws = wb[wb.sheetnames[a]]
            self.new._active_sheet_index = a
            wd = self.new.active
            # print("row:", ws.max_row, "column:", ws.max_column)
            if a == 0:
                ws.merge_cells('D1:H1')
            else:
                ws.merge_cells('A1:I1')
            ws.row_dimensions[1].height = 22
            for i in range(1, ws.max_column + 1):
                column_data = []
                # 遍历列中所有单元格长度
                for ii in range(1, ws.max_row + 1):
                    cell_value = wd.cell(row=ii, column=i).value
                    a_length = len(str(cell_value).encode('utf-8'))
                    # print('a:{},i:{},ii:{},cell_value:{},a_length:{}'.format(a, i, ii, cell_value, a_length))
                    column_data.append(a_length)
                    # ws.row_dimensions[i].height = height
                col_width = max(column_data)
                ws.column_dimensions[get_column_letter(i)].width = col_width
        wb.save(self.create_name)

    def set_row_col(self):
        # 预设特定表格的列宽和行高
        wb = openpyxl.load_workbook(self.create_name)
        for a in range(len(wb.sheetnames)):
            ws = wb[wb.sheetnames[a]]
            if a == 0:
                ws.merge_cells('D1:H1')
                ws.row_dimensions[1].height = 22
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 50
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 11
                ws.column_dimensions['E'].width = 6
                ws.column_dimensions['F'].width = 7
                ws.column_dimensions['G'].width = 13
                ws.column_dimensions['H'].width = 15
            else:
                ws.merge_cells('A1:I1')
                ws.row_dimensions[1].height = 22
                ws.column_dimensions['A'].width = 22
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 11
                ws.column_dimensions['D'].width = 29
                ws.column_dimensions['E'].width = 11
                ws.column_dimensions['F'].width = 10
                ws.column_dimensions['G'].width = 23
                ws.column_dimensions['H'].width = 17
                ws.column_dimensions['I'].width = 9
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 5
        wb.save(self.create_name)


if __name__ == '__main__':
    excel = Excel()
    excel.index_create(sheet_title='目 录', content_title=index_name, content=index_list1)
    excel.excel_create(content_title=excel_name, content=excel_list)
    excel.auto_set_row_col()
    # excel.set_row_col()
